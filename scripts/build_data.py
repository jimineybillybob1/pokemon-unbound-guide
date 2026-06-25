from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


ROOT_DIR = Path(__file__).resolve().parents[1]
DOCS_DIR = ROOT_DIR / "documentation"
FETCHED_REPO_DIR = DOCS_DIR / "fetched-from-repo"
DATA_DIR = ROOT_DIR / "data"
ASSETS_DIR = ROOT_DIR / "assets"
POKEMON_ASSETS_DIR = ASSETS_DIR / "pokemon"
ITEM_ASSETS_DIR = ASSETS_DIR / "items"
UNBOUND_POKEDEX_ENCOUNTERS_FILE = FETCHED_REPO_DIR / "encounters.json"
UNBOUND_POKEDEX_WILD_SOURCE = "Unbound-Pokedex encounters.json"
FETCHED_MOVES_FILE = FETCHED_REPO_DIR / "moves.txt"
FETCHED_MOVE_NAMES_FILE = FETCHED_REPO_DIR / "move_names.txt"
FETCHED_MOVE_DESCRIPTIONS_FILE = FETCHED_REPO_DIR / "move_descriptions.txt"
FETCHED_ABILITIES_FILE = FETCHED_REPO_DIR / "abilities.txt"
FETCHED_ABILITY_NAMES_FILE = FETCHED_REPO_DIR / "ability_names.txt"
FETCHED_ABILITY_DESCRIPTIONS_FILE = FETCHED_REPO_DIR / "ability_descriptions.txt"
FETCHED_DUPLICATE_ABILITIES_FILE = FETCHED_REPO_DIR / "duplicate_abilities.json"

BASE_STATS_FILE = DOCS_DIR / "Base_Stats.txt"
EVOLUTION_FILE = DOCS_DIR / "Evolution Table.txt"
LEARNSETS_FILE = DOCS_DIR / "Learnsets.txt"
EGG_MOVES_FILE = DOCS_DIR / "Egg_Moves.txt"
POKEMON_SPRITE_FALLBACK = "assets/unbound-mark.png"

SKIP_VALUES = {"", "x", "-", "n/a", "none"}
METHOD_MARKERS = {
    "Surfing",
    "Fishing",
    "Old Rod",
    "Good Rod",
    "Super Rod",
    "Rock Smash",
}
UNBOUND_POKEDEX_METHOD_LABELS = {
    "grassAnytime": "Grass",
    "grassDay": "Grass (Day)",
    "grassNight": "Grass (Night)",
    "water": "Surfing",
    "fishing": "Fishing",
    "rockSmash": "Rock Smash",
}
MOVE_CATEGORY_LABELS = {
    "SPLIT_PHYSICAL": "Physical",
    "SPLIT_SPECIAL": "Special",
    "SPLIT_STATUS": "Status",
}
ABILITY_DESCRIPTION_OVERRIDES = {
    "ABILITY_NEUTRALIZINGGAS": "All Abilities are nullified.",
    "ABILITY_FULLMETALBODY": "Prevents ability reduction.",
    "ABILITY_EVAPORATE": "Nullifies all water to up Sp. Atk.",
    "ABILITY_GRASS_DASH": "Grass-type moves hit first.",
    "ABILITY_SLIPPERY_TAIL": "Tail moves hit first.",
    "ABILITY_DRILL_BEAK": "Drill moves land critical hits.",
}

SPECIAL_SPECIES_NAMES = {
    "NIDORAN_F": "Nidoran F",
    "NIDORAN_M": "Nidoran M",
    "MR_MIME": "Mr. Mime",
    "MIME_JR": "Mime Jr.",
    "HO_OH": "Ho-Oh",
    "PORYGON_Z": "Porygon-Z",
    "TYPE_NULL": "Type: Null",
    "JANGMO_O": "Jangmo-o",
    "HAKAMO_O": "Hakamo-o",
    "KOMMO_O": "Kommo-o",
    "FARFETCHD": "Farfetch'd",
    "SIRFETCHD": "Sirfetch'd",
    "FLABEBE": "Flabebe",
    "WISHIWASHI_SCHOOL": "Wishiwashi School",
    "AEGISLASH_BLADE": "Aegislash Blade",
    "ZYGARDE_10": "Zygarde 10%",
    "ZYGARDE_50": "Zygarde 50%",
    "ZYGARDE_COMPLETE": "Zygarde Complete",
}

SPECIAL_SPRITE_SLUGS = {
    "NIDORAN_F": "nidoran-f",
    "NIDORAN_M": "nidoran-m",
    "MR_MIME": "mr-mime",
    "MIME_JR": "mime-jr",
    "HO_OH": "ho-oh",
    "PORYGON_Z": "porygon-z",
    "TYPE_NULL": "type-null",
    "JANGMO_O": "jangmo-o",
    "HAKAMO_O": "hakamo-o",
    "KOMMO_O": "kommo-o",
    "FARFETCHD": "farfetchd",
    "SIRFETCHD": "sirfetchd",
    "WISHIWASHI_S": "wishiwashi-school",
    "AEGISLASH_BLADE": "aegislash-blade",
    "ZYGARDE_10": "zygarde-10",
    "ZYGARDE_50": "zygarde",
    "ZYGARDE_COMPLETE": "zygarde-complete",
    "ORICORIO_Y": "oricorio-pom-pom",
    "ORICORIO_P": "oricorio-pau",
    "ORICORIO_S": "oricorio-sensu",
}

REGIONAL_SUFFIXES = {
    "A": "alola",
    "ALOLA": "alola",
    "G": "galar",
    "GALAR": "galar",
    "H": "hisui",
    "HISUI": "hisui",
    "P": "paldea",
    "PALDEA": "paldea",
}

MOVE_NAME_OVERRIDES = {
    "POISONPOWDER": "Poison Powder",
    "SLEEPPOWDER": "Sleep Powder",
    "STUNSPORE": "Stun Spore",
    "VINEWHIP": "Vine Whip",
    "RAZORLEAF": "Razor Leaf",
    "LEECHSEED": "Leech Seed",
    "TAKEDOWN": "Take Down",
    "SEEDBOMB": "Seed Bomb",
    "SWEETSCENT": "Sweet Scent",
    "WORRYSEED": "Worry Seed",
    "SLUDGEBOMB": "Sludge Bomb",
    "DOUBLEEDGE": "Double-Edge",
    "SOLARBEAM": "Solar Beam",
    "PETALDANCE": "Petal Dance",
    "PETALBLIZZARD": "Petal Blizzard",
    "DRAGONBREATH": "Dragon Breath",
    "FIREFANG": "Fire Fang",
    "FIRESPIN": "Fire Spin",
    "FLAMEBURST": "Flame Burst",
    "SCARYFACE": "Scary Face",
    "FLAREBLITZ": "Flare Blitz",
    "SKULLBASH": "Skull Bash",
    "MAGICALLEAF": "Magical Leaf",
    "GRASSWHISTLE": "Grass Whistle",
    "NATUREPOWER": "Nature Power",
    "LEAFSTORM": "Leaf Storm",
    "POWERWHIP": "Power Whip",
    "GIGADRAIN": "Giga Drain",
    "GRASSYTERRAIN": "Grassy Terrain",
    "BELLYDRUM": "Belly Drum",
    "ANCIENTPOWER": "Ancient Power",
    "DRAGONDANCE": "Dragon Dance",
    "DRAGONRUSH": "Dragon Rush",
    "METALCLAW": "Metal Claw",
    "DRAGONPULSE": "Dragon Pulse",
    "FOCUSPUNCH": "Focus Punch",
    "AIRCUTTER": "Air Cutter",
    "WINGATTACK": "Wing Attack",
    "DRAGONTAIL": "Dragon Tail",
    "MIRRORCOAT": "Mirror Coat",
    "MUDSPORT": "Mud Sport",
    "MUDDYWATER": "Muddy Water",
    "FAKEOUT": "Fake Out",
    "AQUARING": "Aqua Ring",
    "AQUAJET": "Aqua Jet",
    "WATERSPOUT": "Water Spout",
    "LIFEDEW": "Life Dew",
    "FOLLOWME": "Follow Me",
}


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return re.sub(r"\s+", " ", text)


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def read_optional_text(path: Path) -> str:
    if not path.exists():
        return ""
    return read_text(path)


def read_json_file(path: Path) -> object | None:
    if not path.exists():
        return None
    try:
        return json.loads(read_text(path))
    except json.JSONDecodeError:
        return None


def normalize_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())


def parse_source_string(value: str) -> str:
    text = clean(value)
    if not text:
        return ""
    text = text.replace('\\"', '"')
    text = text.replace("\\n", " ")
    text = text.replace("\\p", " ")
    text = text.replace("\\l", " ")
    text = re.sub(r"\\([A-Za-z])", r"\1", text)
    text = re.sub(r"\s+", " ", text).strip()
    return "" if text == "-" else text


def strip_source_conditionals(text: str) -> str:
    output: list[str] = []
    include_stack: list[bool] = []
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith("#ifdef") or stripped.startswith("#ifndef"):
            include_stack.append(True)
            continue
        if stripped.startswith("#else"):
            if include_stack:
                include_stack[-1] = not include_stack[-1]
            continue
        if stripped.startswith("#endif"):
            if include_stack:
                include_stack.pop()
            continue
        if stripped.startswith("#include"):
            continue
        if not include_stack or all(include_stack):
            output.append(line)
    return "\n".join(output)


def parse_org_entries(text: str, prefix: str) -> dict[str, str]:
    entries: dict[str, str] = {}
    labels: list[str] = []
    buffer: list[str] = []

    def flush() -> None:
        if not labels:
            return
        content = parse_source_string(" ".join(part.strip() for part in buffer if clean(part)))
        if content:
            for label in labels:
                key = label.removeprefix(prefix) if label.startswith(prefix) else label
                entries[normalize_key(key)] = content

    for line in strip_source_conditionals(text).splitlines():
        stripped = line.strip()
        if stripped.startswith("#org @"):
            if labels and buffer:
                flush()
                labels = []
                buffer = []
            labels.append(stripped.removeprefix("#org @"))
            continue
        if labels:
            buffer.append(line)

    if labels:
        flush()
    return entries


def load_reference_metadata() -> dict:
    metadata = {
        "moves": {},
        "moveNames": {},
        "abilities": {},
        "sources": {},
        "warnings": [],
    }

    moves_text = read_optional_text(FETCHED_MOVES_FILE)
    move_names_text = read_optional_text(FETCHED_MOVE_NAMES_FILE)
    move_descriptions_text = read_optional_text(FETCHED_MOVE_DESCRIPTIONS_FILE)
    if moves_text and move_names_text and move_descriptions_text:
        move_names = parse_org_entries(move_names_text, "NAME_")
        move_descriptions = parse_org_entries(move_descriptions_text, "DESC_")
        field_re = re.compile(r"\.(\w+)\s*=\s*([^,\n]+),")
        markers = list(re.finditer(r"\[MOVE_([A-Z0-9_]+)\]\s*=", moves_text))
        for index, match in enumerate(markers):
            move_id = match.group(1)
            if move_id == "NONE":
                continue
            start = match.end()
            end = markers[index + 1].start() if index + 1 < len(markers) else len(moves_text)
            fields = {field: clean(value) for field, value in field_re.findall(moves_text[start:end])}
            move_name = move_names.get(normalize_key(move_id)) or MOVE_NAME_OVERRIDES.get(move_id) or token_words(move_id)
            metadata["moveNames"][normalize_key(move_id)] = move_name
            metadata["moves"][normalize_key(move_name)] = {
                "type": humanize_prefixed(fields.get("type", ""), "TYPE_"),
                "category": MOVE_CATEGORY_LABELS.get(fields.get("split", ""), ""),
                "power": parse_int(fields.get("power", "")),
                "accuracy": parse_int(fields.get("accuracy", "")),
                "pp": parse_int(fields.get("pp", "")),
                "priority": parse_int(fields.get("priority", "")),
                "contact": "FLAG_MAKES_CONTACT" in fields.get("flags", ""),
                "description": move_descriptions.get(normalize_key(move_id), ""),
                "effect": "",
                "metadataSource": "Fetched CFRU move data",
            }
        metadata["sources"]["moves"] = {
            "name": "Fetched CFRU move data",
            "files": [
                FETCHED_MOVES_FILE.name,
                FETCHED_MOVE_NAMES_FILE.name,
                FETCHED_MOVE_DESCRIPTIONS_FILE.name,
            ],
        }
    else:
        metadata["warnings"].append(
            "Move metadata requires fetched-from-repo move source files. Run scripts\\fetch_unbound_source.py to populate them."
        )

    abilities_text = read_optional_text(FETCHED_ABILITIES_FILE)
    ability_names_text = read_optional_text(FETCHED_ABILITY_NAMES_FILE)
    ability_descriptions_text = read_optional_text(FETCHED_ABILITY_DESCRIPTIONS_FILE)
    duplicate_abilities_payload = read_json_file(FETCHED_DUPLICATE_ABILITIES_FILE)
    if abilities_text and ability_names_text and ability_descriptions_text:
        ability_names = parse_org_entries(ability_names_text, "NAME_")
        ability_descriptions = parse_org_entries(ability_descriptions_text, "DESC_")
        ability_by_constant: dict[str, dict] = {}
        for match in re.finditer(r"#define\s+(ABILITY_[A-Z0-9_]+)\s+(\d+)", abilities_text):
            constant = match.group(1)
            if constant == "ABILITY_NONE":
                continue
            ability_id = constant.removeprefix("ABILITY_")
            name = ability_names.get(normalize_key(ability_id)) or token_words(ability_id)
            description = ability_descriptions.get(normalize_key(ability_id), "")
            record = {
                "name": name,
                "description": description,
                "metadataSource": "Fetched CFRU ability data" if description else "",
            }
            ability_by_constant[constant] = record
            if description:
                metadata["abilities"][normalize_key(name)] = record

        if isinstance(duplicate_abilities_payload, dict):
            for base_constant, variants in duplicate_abilities_payload.items():
                base_record = ability_by_constant.get(base_constant)
                base_description = base_record.get("description", "") if base_record else ""
                for variant_constant in variants or {}:
                    variant_record = ability_by_constant.get(variant_constant)
                    if not variant_record:
                        variant_name = humanize_prefixed(variant_constant, "ABILITY_")
                        variant_record = {"name": variant_name, "description": "", "metadataSource": ""}
                        ability_by_constant[variant_constant] = variant_record
                    if base_description and not variant_record.get("description"):
                        variant_record["description"] = base_description
                        variant_record["metadataSource"] = "Unbound-Pokedex duplicate ability mapping"
                        metadata["abilities"][normalize_key(variant_record["name"])] = variant_record

        for constant, description in ABILITY_DESCRIPTION_OVERRIDES.items():
            record = ability_by_constant.get(constant)
            if not record:
                continue
            record["description"] = description
            record["metadataSource"] = "Unbound-Pokedex ability override"
            metadata["abilities"][normalize_key(record["name"])] = record

        metadata["sources"]["abilities"] = {
            "name": "Fetched CFRU and Unbound-Pokedex ability data",
            "files": [
                FETCHED_ABILITIES_FILE.name,
                FETCHED_ABILITY_NAMES_FILE.name,
                FETCHED_ABILITY_DESCRIPTIONS_FILE.name,
                FETCHED_DUPLICATE_ABILITIES_FILE.name,
            ],
        }
    else:
        metadata["warnings"].append(
            "Ability metadata requires fetched-from-repo ability source files. Run scripts\\fetch_unbound_source.py to populate them."
        )

    return metadata


def asset_slug(value: object) -> str:
    text = clean(value).lower()
    text = (
        text.replace("'", "")
        .replace(":", "")
        .replace(".", "")
        .replace("%", "")
        .replace("♀", "f")
        .replace("♂", "m")
    )
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def relative_asset(path: Path) -> str:
    return path.relative_to(ROOT_DIR).as_posix()


def build_asset_index(root: Path) -> dict[str, list[str]]:
    index: dict[str, list[str]] = defaultdict(list)
    if not root.exists():
        return index

    for path in sorted(root.rglob("*.png")):
        rel = relative_asset(path)
        stem = path.stem.lower()
        keys = {stem}
        if stem.endswith("--bag"):
            keys.add(stem.removesuffix("--bag"))
        if stem.endswith("--held"):
            keys.add(stem.removesuffix("--held"))
        for key in keys:
            index[key].append(rel)

    for key, paths in index.items():
        paths.sort(
            key=lambda item: (
                0 if "/custom/" in item else 1,
                0 if item.endswith("--bag.png") else 1,
                1 if item.endswith("--held.png") else 0,
                item.count("/"),
                item,
            )
        )
    return index


def sprite_base_slug(species_id: str) -> str:
    if species_id in SPECIAL_SPRITE_SLUGS:
        return SPECIAL_SPRITE_SLUGS[species_id]
    return asset_slug(token_words(species_id))


def species_sprite_candidates(constant: str, display_name: str) -> list[str]:
    species_id = constant.removeprefix("SPECIES_")
    tokens = species_id.split("_")
    candidates: list[str] = []

    def add(candidate: str) -> None:
        if candidate and candidate not in candidates:
            candidates.append(candidate)

    add(SPECIAL_SPRITE_SLUGS.get(species_id, ""))

    if tokens and tokens[-1] in REGIONAL_SUFFIXES:
        region = REGIONAL_SUFFIXES[tokens[-1]]
        base_id = "_".join(tokens[:-1])
        add(f"{sprite_base_slug(base_id)}-{region}")

    if tokens and tokens[-1] == "MEGA":
        add(f"{sprite_base_slug('_'.join(tokens[:-1]))}-mega")
    elif len(tokens) >= 2 and tokens[-2] == "MEGA":
        add(f"{sprite_base_slug('_'.join(tokens[:-2]))}-mega-{asset_slug(tokens[-1])}")

    if tokens and tokens[-1] == "GIGA":
        add(f"{sprite_base_slug('_'.join(tokens[:-1]))}-gmax")
    if tokens and tokens[-1] == "PRIMAL":
        add(f"{sprite_base_slug('_'.join(tokens[:-1]))}-primal")

    if len(tokens) > 1:
        add(f"{sprite_base_slug(tokens[0])}-{'-'.join(asset_slug(part) for part in tokens[1:] if part)}")

    add(sprite_base_slug(species_id))
    add(asset_slug(display_name))
    return candidates


def resolve_sprite(pokemon_entry: dict, sprite_index: dict[str, list[str]]) -> str:
    for candidate in species_sprite_candidates(pokemon_entry["constant"], pokemon_entry["name"]):
        matches = sprite_index.get(candidate)
        if matches:
            return matches[0]
    return POKEMON_SPRITE_FALLBACK


def resolve_item_icon(
    name: str,
    item_index: dict[str, list[str]],
    *,
    item_type: str = "",
    machine_kind: str = "",
) -> str:
    slug = asset_slug(name)
    candidates = [slug, f"{slug}--bag", f"{slug}--held"]
    if machine_kind and item_type:
        candidates.append(f"{asset_slug(item_type)}")
    if machine_kind == "tm":
        candidates.extend(["tm", "normal"])
    elif machine_kind == "hm":
        candidates.extend(["hm", "normal"])

    for candidate in candidates:
        matches = item_index.get(candidate)
        if not matches:
            continue
        if machine_kind and item_type:
            typed_path = f"assets/items/{machine_kind}/{asset_slug(item_type)}.png"
            if typed_path in matches:
                return typed_path
        return matches[0]
    return ""


def token_words(token: str) -> str:
    token = token.strip("_")
    if not token:
        return ""
    if token in {"HP", "ID", "IV", "EV", "BP", "TM", "HM"}:
        return token
    return " ".join(part.capitalize() for part in token.split("_") if part)


def humanize_map_name(value: str) -> str:
    words = []
    for token in clean(value).split("_"):
        if not token:
            continue
        if any(char.isdigit() for char in token):
            words.append(token.upper())
        elif len(token) == 1 and token.isalpha():
            words.append(token.upper())
        else:
            words.append(token.capitalize())
    return " ".join(words)


def humanize_prefixed(value: str, prefix: str) -> str:
    if not value or value.endswith("_NONE"):
        return ""
    body = value[len(prefix) :] if value.startswith(prefix) else value
    return token_words(body)


def humanize_species(value: str) -> str:
    species_id = value[len("SPECIES_") :] if value.startswith("SPECIES_") else value
    if species_id in SPECIAL_SPECIES_NAMES:
        return SPECIAL_SPECIES_NAMES[species_id]

    tokens = species_id.split("_")
    prefix = ""
    suffix = ""

    if tokens and tokens[-1] in {"A", "ALOLA"}:
        prefix = "Alolan "
        tokens = tokens[:-1]
    elif tokens and tokens[-1] in {"G", "GALAR"}:
        prefix = "Galarian "
        tokens = tokens[:-1]
    elif tokens and tokens[-1] in {"H", "HISUI"}:
        prefix = "Hisuian "
        tokens = tokens[:-1]
    elif tokens and tokens[-1] in {"P", "PALDEA"}:
        prefix = "Paldean "
        tokens = tokens[:-1]

    if tokens and tokens[-1] == "MEGA":
        prefix = "Mega " + prefix
        tokens = tokens[:-1]
    elif len(tokens) >= 2 and tokens[-2] == "MEGA":
        prefix = "Mega " + prefix
        suffix = f" {tokens[-1]}"
        tokens = tokens[:-2]
    elif tokens and tokens[-1] == "GIGA":
        prefix = "Gigantamax "
        tokens = tokens[:-1]
    elif tokens and tokens[-1] == "PRIMAL":
        prefix = "Primal "
        tokens = tokens[:-1]

    base_key = "_".join(tokens)
    base_name = SPECIAL_SPECIES_NAMES.get(base_key, token_words(base_key))
    return f"{prefix}{base_name}{suffix}".strip()


def move_name_from_constant(
    move_constant: str,
    hints: dict[str, str] | None = None,
    name_lookup: dict[str, str] | None = None,
) -> str:
    body = move_constant[len("MOVE_") :] if move_constant.startswith("MOVE_") else move_constant
    if body in MOVE_NAME_OVERRIDES:
        return MOVE_NAME_OVERRIDES[body]
    if name_lookup:
        named = name_lookup.get(normalize_key(body))
        if named:
            return named
    if hints:
        hinted = hints.get(normalize_key(body))
        if hinted:
            return hinted
    return token_words(body)


def parse_int(value: str) -> int | None:
    value = clean(value)
    if not value:
        return None
    match = re.search(r"-?\d+", value)
    return int(match.group(0)) if match else None


def find_workbook(fragment: str) -> Path:
    matches = sorted(DOCS_DIR.glob(f"*{fragment}*.xlsx"))
    if not matches:
        raise FileNotFoundError(f"No workbook matching *{fragment}*.xlsx")
    return matches[0]


def rows_for(workbook: openpyxl.Workbook, sheet_name: str) -> list[list[str]]:
    ws = workbook[sheet_name]
    return [[clean(cell) for cell in row] for row in ws.iter_rows(values_only=True)]


def compact_row(row: list[str]) -> list[str]:
    return [value for value in row if clean(value)]


def is_data_value(value: str) -> bool:
    return normalize_key(value) not in SKIP_VALUES


def split_top_level_commas(value: str) -> list[str]:
    parts: list[str] = []
    current: list[str] = []
    depth = 0
    for char in value:
        if char == "(":
            depth += 1
        elif char == ")" and depth:
            depth -= 1
        if char == "," and depth == 0:
            parts.append("".join(current).strip())
            current = []
        else:
            current.append(char)
    if current:
        parts.append("".join(current).strip())
    return parts


def parse_base_stats() -> dict[str, dict]:
    text = read_text(BASE_STATS_FILE)
    field_re = re.compile(r"\.(\w+)\s*=\s*([^,\n]+),")
    pokemon: dict[str, dict] = {}

    markers = list(re.finditer(r"\[SPECIES_([A-Z0-9_]+)\]\s*=", text))
    for index, match in enumerate(markers):
        species_id = match.group(1)
        if species_id == "NONE":
            continue
        start = match.end()
        end = markers[index + 1].start() if index + 1 < len(markers) else len(text)
        fields = {field: clean(value) for field, value in field_re.findall(text[start:end])}
        if not fields:
            continue

        constant = f"SPECIES_{species_id}"
        stats = {
            "hp": parse_int(fields.get("baseHP", "")) or 0,
            "atk": parse_int(fields.get("baseAttack", "")) or 0,
            "def": parse_int(fields.get("baseDefense", "")) or 0,
            "spa": parse_int(fields.get("baseSpAttack", "")) or 0,
            "spdef": parse_int(fields.get("baseSpDefense", "")) or 0,
            "spd": parse_int(fields.get("baseSpeed", "")) or 0,
        }
        type_constants = [fields.get("type1", ""), fields.get("type2", "")]
        types = []
        for type_constant in type_constants:
            label = humanize_prefixed(type_constant, "TYPE_")
            if label and label not in types:
                types.append(label)

        ability_constants = [fields.get("ability1", ""), fields.get("ability2", "")]
        abilities = []
        for ability_constant in ability_constants:
            label = humanize_prefixed(ability_constant, "ABILITY_")
            if label and label not in abilities:
                abilities.append(label)

        hidden_ability = humanize_prefixed(fields.get("hiddenAbility", ""), "ABILITY_")
        held_items = []
        for slot, field_name in (("common", "item1"), ("rare", "item2")):
            item = humanize_prefixed(fields.get(field_name, ""), "ITEM_")
            if item:
                held_items.append({"slot": slot, "item": item})

        pokemon[constant] = {
            "id": normalize_key(species_id),
            "constant": constant,
            "name": humanize_species(constant),
            "isPlaceholder": constant in {"SPECIES_EGG", "SPECIES_MANAPHY_EGG"},
            "stats": stats,
            "bst": sum(stats.values()),
            "types": types,
            "abilities": abilities,
            "hiddenAbility": hidden_ability,
            "heldItems": held_items,
            "catchRate": parse_int(fields.get("catchRate", "")),
            "expYield": parse_int(fields.get("expYield", "")),
            "growthRate": humanize_prefixed(fields.get("growthRate", ""), "GROWTH_"),
            "eggGroups": [
                group
                for group in (
                    humanize_prefixed(fields.get("eggGroup1", ""), "EGG_GROUP_"),
                    humanize_prefixed(fields.get("eggGroup2", ""), "EGG_GROUP_"),
                )
                if group
            ],
            "evYield": {
                "hp": parse_int(fields.get("evYield_HP", "")) or 0,
                "atk": parse_int(fields.get("evYield_Attack", "")) or 0,
                "def": parse_int(fields.get("evYield_Defense", "")) or 0,
                "spa": parse_int(fields.get("evYield_SpAttack", "")) or 0,
                "spdef": parse_int(fields.get("evYield_SpDefense", "")) or 0,
                "spd": parse_int(fields.get("evYield_Speed", "")) or 0,
            },
            "levelUpLearnset": [],
            "eggMoves": [],
            "evolutions": [],
            "locations": [],
        }

    return pokemon


def parse_learnsets() -> dict[str, list[dict]]:
    text = read_text(LEARNSETS_FILE)
    table_re = re.compile(
        r"static const struct LevelUpMove\s+(s\w+LevelUpLearnset)\[\]\s*=\s*\{(.*?)\n\};",
        re.S,
    )
    move_re = re.compile(r"LEVEL_UP_MOVE\(\s*(\d+)\s*,\s*(MOVE_[A-Z0-9_]+)\s*\)")
    mapping_re = re.compile(r"\[SPECIES_([A-Z0-9_]+)\]\s*=\s*(s\w+LevelUpLearnset|sEmptyMoveset),")

    tables = {
        name: [
            {"level": int(level), "moveConstant": move_constant}
            for level, move_constant in move_re.findall(body)
        ]
        for name, body in table_re.findall(text)
    }

    learnsets: dict[str, list[dict]] = {}
    for species_id, table_name in mapping_re.findall(text):
        constant = f"SPECIES_{species_id}"
        learnsets[constant] = list(tables.get(table_name, []))
    return learnsets


def parse_egg_moves() -> dict[str, list[str]]:
    text = read_text(EGG_MOVES_FILE)
    egg_moves: dict[str, list[str]] = {}
    for match in re.finditer(r"egg_moves\(\s*([A-Z0-9_]+)\s*,(.*?)\)", text, re.S):
        species = f"SPECIES_{match.group(1)}"
        moves = re.findall(r"MOVE_[A-Z0-9_]+", match.group(2))
        egg_moves[species] = moves
    return egg_moves


def evo_entry_bodies(chunk: str) -> list[str]:
    chunk = re.sub(r"//.*", "", chunk)
    bodies: list[str] = []
    depth = 0
    current: list[str] | None = None
    for char in chunk:
        if char == "{":
            depth += 1
            if depth == 2:
                current = []
                continue
        elif char == "}":
            if depth == 2 and current is not None:
                bodies.append("".join(current).strip())
                current = None
            depth -= 1
            continue
        if current is not None:
            current.append(char)
    return [body for body in bodies if body and body != "0"]


def evolution_label(method: str, args: list[str], target: str) -> str:
    arg0 = args[0] if args else ""
    arg1 = args[1] if len(args) > 1 else ""
    if method == "EVO_LEVEL":
        return f"Level {arg0}"
    if method == "EVO_LEVEL_ATK_GT_DEF":
        return f"Level {arg0}, Attack > Defense"
    if method == "EVO_LEVEL_ATK_EQ_DEF":
        return f"Level {arg0}, Attack = Defense"
    if method == "EVO_LEVEL_ATK_LT_DEF":
        return f"Level {arg0}, Attack < Defense"
    if method == "EVO_LEVEL_FEMALE":
        return f"Level {arg0}, female"
    if method == "EVO_LEVEL_MALE":
        return f"Level {arg0}, male"
    if method == "EVO_LEVEL_DAY":
        return f"Level {arg0} during the day"
    if method == "EVO_LEVEL_NIGHT":
        return f"Level {arg0} at night"
    if method == "EVO_ITEM":
        return f"Use {humanize_prefixed(arg0, 'ITEM_')}"
    if method == "EVO_ITEM_MALE":
        return f"Use {humanize_prefixed(arg0, 'ITEM_')} on a male Pokemon"
    if method == "EVO_ITEM_FEMALE":
        return f"Use {humanize_prefixed(arg0, 'ITEM_')} on a female Pokemon"
    if method == "EVO_ITEM_LOCATION":
        return f"Use {humanize_prefixed(arg0, 'ITEM_')} at {humanize_prefixed(arg1, 'MAPSEC_')}"
    if method == "EVO_TRADE":
        return "Trade"
    if method == "EVO_TRADE_ITEM":
        return f"Trade holding {humanize_prefixed(arg0, 'ITEM_')}"
    if method == "EVO_FRIENDSHIP":
        return "High friendship"
    if method == "EVO_FRIENDSHIP_DAY":
        return "High friendship during the day"
    if method == "EVO_FRIENDSHIP_NIGHT":
        return "High friendship at night"
    if method == "EVO_MAP":
        return f"Level up at {humanize_prefixed(arg0, 'MAPSEC_')}"
    if method == "EVO_MOVE":
        return f"Level up knowing {move_name_from_constant(arg0)}"
    if method == "EVO_MEGA":
        return f"Mega Evolution with {humanize_prefixed(arg0, 'ITEM_')}"
    if method == "EVO_GIGANTAMAX":
        return "Gigantamax form"
    return token_words(method.removeprefix("EVO_"))


def parse_evolutions() -> dict[str, list[dict]]:
    text = read_text(EVOLUTION_FILE)
    markers = list(re.finditer(r"\[SPECIES_([A-Z0-9_]+)\]\s*=", text))
    evolutions: dict[str, list[dict]] = {}

    for index, marker in enumerate(markers):
        species = f"SPECIES_{marker.group(1)}"
        start = marker.end()
        end = markers[index + 1].start() if index + 1 < len(markers) else len(text)
        chunk = text[start:end]
        entries = []
        for body in evo_entry_bodies(chunk):
            parts = split_top_level_commas(body)
            if not parts or not parts[0].startswith("EVO_"):
                continue
            target = next((part for part in parts if part.startswith("SPECIES_")), "")
            if not target:
                continue
            target_index = parts.index(target)
            args = parts[1:target_index]
            method = parts[0]
            entries.append(
                {
                    "methodId": method,
                    "method": evolution_label(method, args, target),
                    "target": target,
                    "targetName": humanize_species(target),
                    "args": args,
                    "raw": body,
                }
            )
        if entries:
            evolutions[species] = entries

    return evolutions


def parse_simple_table(rows: list[list[str]], start_row: int = 1) -> list[dict]:
    headers = compact_row(rows[start_row - 1])
    records: list[dict] = []
    for row in rows[start_row:]:
        values = compact_row(row)
        if not values:
            continue
        record = {}
        for index, header in enumerate(headers):
            record[normalize_key(header) or f"col{index + 1}"] = values[index] if index < len(values) else ""
        records.append(record)
    return records


def add_location_method(
    locations: dict[str, dict], location_name: str, method_label: str, species: list[str], source: str
) -> None:
    location_name = clean(location_name)
    species = [clean(name) for name in species if is_data_value(name) and clean(name) not in METHOD_MARKERS]
    if not location_name or not species:
        return
    key = normalize_key(location_name)
    location = locations.setdefault(
        key,
        {"id": key, "name": location_name, "methods": [], "special": []},
    )
    existing = next((method for method in location["methods"] if method["label"] == method_label), None)
    if not existing:
        existing = {"label": method_label, "species": [], "source": source}
        location["methods"].append(existing)
    for name in species:
        if name not in existing["species"]:
            existing["species"].append(name)


def parse_grass_cave(rows: list[list[str]], locations: dict[str, dict]) -> None:
    headers = rows[0]
    for col, location_name in enumerate(headers):
        if not clean(location_name):
            continue
        species = [row[col] for row in rows[1:] if col < len(row)]
        add_location_method(locations, location_name, "Grass / cave", species, "Grass & Cave Encounters")


def parse_surfing_fishing(rows: list[list[str]], locations: dict[str, dict]) -> None:
    location_row = rows[2]
    location_cols = [(col, name) for col, name in enumerate(location_row) if clean(name)]
    current_section = ""
    current_method_by_col: dict[int, str] = {}

    for row_index, row in enumerate(rows):
        row_values = compact_row(row)
        if len(row_values) == 1 and row_values[0] in {"Surfing", "Fishing"}:
            current_section = row_values[0]
            current_method_by_col.clear()
            continue
        if row_index <= 2:
            continue

        for col, location_name in location_cols:
            value = row[col] if col < len(row) else ""
            if not is_data_value(value):
                continue
            if value in METHOD_MARKERS:
                if value in {"Old Rod", "Good Rod", "Super Rod", "Rock Smash"}:
                    current_method_by_col[col] = value
                continue
            method = current_method_by_col.get(col) or current_section or "Water encounter"
            add_location_method(locations, location_name, method, [value], "Surfing, Fishing, Rock Smash")


def parse_gift_static(rows: list[list[str]]) -> list[dict]:
    records = []
    for row in rows[3:]:
        method, location, pokemon, requirement = (row + ["", "", "", ""])[:4]
        if not any((method, location, pokemon, requirement)):
            continue
        records.append(
            {
                "method": method,
                "location": location,
                "pokemon": pokemon,
                "requirement": requirement,
            }
        )
    return records


def parse_legendary(rows: list[list[str]]) -> list[dict]:
    records: list[dict] = []
    header_rows = []
    for index in range(len(rows) - 1):
        next_values = [clean(value) for value in rows[index + 1][:4]]
        if next_values[:2] == ["Requires", "Location"]:
            header_rows.append(index)

    for idx, header_row in enumerate(header_rows):
        end = header_rows[idx + 1] if idx + 1 < len(header_rows) else len(rows)
        header = rows[header_row]
        for col in range(0, len(header), 2):
            name = clean(header[col])
            if not name:
                continue
            requirements = []
            locations = []
            notes = []
            for row in rows[header_row + 2 : end]:
                req = row[col] if col < len(row) else ""
                loc = row[col + 1] if col + 1 < len(row) else ""
                if req:
                    requirements.append(req)
                if loc:
                    locations.append(loc)
                if not req and not loc and any(compact_row(row)):
                    continue
            if requirements or locations:
                records.append(
                    {
                        "pokemon": name,
                        "requirements": requirements,
                        "locations": locations,
                        "notes": notes,
                    }
                )
    return records


def parse_trades_game_corner(rows: list[list[str]]) -> dict:
    trades = []
    game_corner = []
    group_cols = [0, 5, 10, 15, 20, 25]
    for col in group_cols:
        location = rows[3][col] if col < len(rows[3]) else ""
        if not location:
            continue
        for row_index in (5, 6):
            send = rows[row_index][col] if col < len(rows[row_index]) else ""
            get = rows[row_index][col + 2] if col + 2 < len(rows[row_index]) else ""
            if send and get:
                trades.append({"location": location, "send": send, "receive": get})

    for col in group_cols:
        coins = rows[10][col] if col < len(rows[10]) else ""
        pokemon = rows[11][col] if col < len(rows[11]) else ""
        if coins and pokemon:
            game_corner.append({"location": "Dehara Game Corner", "coins": parse_int(coins), "pokemon": pokemon})

    raw_sections = []
    for row in rows[13:]:
        values = compact_row(row)
        if values:
            raw_sections.append(values)

    return {"trades": trades, "gameCorner": game_corner, "rawSections": raw_sections}


def parse_swarm(rows: list[list[str]]) -> list[dict]:
    times = rows[0][1:]
    schedule = []
    for row in rows[1:]:
        day = row[0] if row else ""
        if not day:
            continue
        entries = []
        for time, pokemon in zip(times, row[1:]):
            if time and pokemon:
                entries.append({"time": time, "pokemon": pokemon})
        if entries:
            schedule.append({"day": parse_int(day) or day, "entries": entries})
    return schedule


def parse_wild_held_items(rows: list[list[str]]) -> list[dict]:
    records = []
    for row in rows[1:]:
        number = row[0] if len(row) > 0 else ""
        pokemon = row[2] if len(row) > 2 else ""
        if not pokemon:
            continue
        item_values = []
        for slot, index in (("50%", 5), ("5%", 7)):
            value = row[index] if index < len(row) else ""
            if value and normalize_key(value) != "x":
                item_values.append({"slot": slot, "item": value})
        records.append(
            {
                "number": number,
                "pokemon": pokemon,
                "items": item_values,
            }
        )
    return records


def parse_tm_hm(rows: list[list[str]]) -> list[dict]:
    records = []
    for row in rows[1:]:
        values = compact_row(row)
        if len(values) < 3:
            continue
        number, name = values[0], values[1]
        if len(values) >= 4:
            type_name = values[2]
            location = values[-1]
        else:
            type_name = ""
            location = values[2]
        records.append({"number": number, "name": name, "type": type_name, "location": location})
    return records


def parse_item_locations(rows: list[list[str]], has_type: bool = False) -> list[dict]:
    records = []
    for row in rows[1:]:
        values = compact_row(row)
        if len(values) < 3:
            continue
        record = {"number": values[0], "name": values[1], "location": values[-1]}
        if has_type and len(values) >= 4:
            record["type"] = values[2]
        records.append(record)
    return records


def parse_frontier_services(rows: list[list[str]]) -> list[dict]:
    services = []
    for col, name in enumerate(rows[0][1:], start=1):
        if not name:
            continue
        offerings = []
        for row in rows[6:]:
            value = row[col] if col < len(row) else ""
            if value:
                offerings.append(value)
        services.append(
            {
                "name": name,
                "location": rows[1][col] if col < len(rows[1]) else "",
                "availability": rows[2][col] if col < len(rows[2]) else "",
                "cost": rows[4][col] if col < len(rows[4]) else "",
                "details": rows[5][col] if col < len(rows[5]) else "",
                "offerings": offerings,
            }
        )
    return services


def split_move_cost(value: str) -> dict:
    if " - " in value:
        move, cost = value.rsplit(" - ", 1)
        return {"move": clean(move), "cost": clean(cost)}
    return {"move": clean(value), "cost": ""}


def parse_move_tutors(rows: list[list[str]]) -> list[dict]:
    tutors = []
    for col, name in enumerate(rows[0][1:], start=1):
        if not name:
            continue
        moves = []
        for row in rows[5:]:
            value = row[col] if col < len(row) else ""
            if value:
                moves.append(split_move_cost(value))
        tutors.append(
            {
                "name": name,
                "location": rows[1][col] if col < len(rows[1]) else "",
                "availability": rows[2][col] if col < len(rows[2]) else "",
                "moves": moves,
            }
        )
    return tutors


def parse_location_workbook() -> tuple[dict, dict[str, str]]:
    path = find_workbook("Location Guide")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    locations: dict[str, dict] = {}
    move_hints: dict[str, str] = {}

    parse_grass_cave(rows_for(workbook, "Grass & Cave Encounters"), locations)
    parse_surfing_fishing(rows_for(workbook, "Surfing, Fishing, Rock Smash"), locations)

    gift_static = parse_gift_static(rows_for(workbook, "Gift & Static Encounters"))
    legendary = parse_legendary(rows_for(workbook, "LegendMythical & Ultra Beasts"))
    trade_data = parse_trades_game_corner(rows_for(workbook, "In-game trades & Game Corner"))
    swarms = parse_swarm(rows_for(workbook, "Swarm Schedule"))
    wild_held_items = parse_wild_held_items(rows_for(workbook, "Wild Held Items"))
    tm_hm = parse_tm_hm(rows_for(workbook, "TM & HM"))
    mega_stones = parse_item_locations(rows_for(workbook, "Megastone"))
    z_crystals = parse_item_locations(rows_for(workbook, "Z-Crystal"), has_type=True)

    for item in tm_hm:
        move_hints[normalize_key(item["name"])] = item["name"]

    workbook.close()
    return (
        {
            "source": path.name,
            "locations": sorted(locations.values(), key=lambda item: item["name"]),
            "giftStatic": gift_static,
            "legendary": legendary,
            "trades": trade_data["trades"],
            "gameCorner": trade_data["gameCorner"],
            "tradeRawSections": trade_data["rawSections"],
            "swarms": swarms,
            "wildHeldItems": wild_held_items,
            "tmHm": tm_hm,
            "megaStones": mega_stones,
            "zCrystals": z_crystals,
        },
        move_hints,
    )


def parse_unbound_pokedex_wild_locations(pokemon: dict[str, dict]) -> list[dict]:
    payload = read_json_file(UNBOUND_POKEDEX_ENCOUNTERS_FILE)
    if not isinstance(payload, list):
        return []

    species_names = {constant: entry["name"] for constant, entry in pokemon.items()}
    locations_by_name: dict[str, dict[str, list[str]]] = {}

    for entry in payload:
        if not isinstance(entry, dict):
            continue
        map_name = humanize_map_name(clean(entry.get("mapName")))
        encounters = entry.get("encounters")
        if not map_name or not isinstance(encounters, dict):
            continue
        method_map = locations_by_name.setdefault(map_name, {})
        for key, label in UNBOUND_POKEDEX_METHOD_LABELS.items():
            details = encounters.get(key)
            if not isinstance(details, dict):
                continue
            slots = details.get("slots")
            if not isinstance(slots, list):
                continue
            species_list = method_map.setdefault(label, [])
            for slot in slots:
                if not isinstance(slot, dict):
                    continue
                species_constant = clean(slot.get("species"))
                if not species_constant:
                    continue
                species_name = species_names.get(species_constant) or humanize_species(species_constant)
                if species_name and species_name not in species_list:
                    species_list.append(species_name)

    records = []
    for location_name in sorted(locations_by_name):
        methods = []
        for label in UNBOUND_POKEDEX_METHOD_LABELS.values():
            species_list = locations_by_name[location_name].get(label, [])
            if species_list:
                methods.append(
                    {
                        "label": label,
                        "species": species_list,
                        "source": UNBOUND_POKEDEX_WILD_SOURCE,
                    }
                )
        if methods:
            records.append({"name": location_name, "methods": methods})
    return records


def parse_frontier_workbook() -> tuple[dict, dict[str, str]]:
    path = find_workbook("Frontier Documentation")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    services = parse_frontier_services(rows_for(workbook, "Battle Frontier Services"))
    tutors = parse_move_tutors(rows_for(workbook, "Move Tutors"))
    daily_events = parse_simple_table(rows_for(workbook, "Daily Events"))
    move_hints = {}
    for tutor in tutors:
        for move in tutor["moves"]:
            move_hints[normalize_key(move["move"])] = move["move"]
    workbook.close()
    return (
        {
            "source": path.name,
            "services": services,
            "moveTutors": tutors,
            "dailyEvents": daily_events,
        },
        move_hints,
    )


def attach_source_data(
    pokemon: dict[str, dict],
    learnsets: dict[str, list[dict]],
    egg_moves: dict[str, list[str]],
    evolutions: dict[str, list[dict]],
) -> None:
    for species, entries in learnsets.items():
        if species in pokemon:
            pokemon[species]["levelUpLearnset"] = entries
    for species, moves in egg_moves.items():
        if species in pokemon:
            pokemon[species]["eggMoves"] = [{"moveConstant": move} for move in moves]
    for species, edges in evolutions.items():
        if species in pokemon:
            pokemon[species]["evolutions"] = edges


def collect_move_data(
    pokemon: dict[str, dict],
    move_hints: dict[str, str],
    move_metadata: dict[str, dict],
    move_name_lookup: dict[str, str],
) -> list[dict]:
    moves: dict[str, dict] = {}

    def ensure_move(move_constant: str) -> dict:
        name = move_name_from_constant(move_constant, move_hints, move_name_lookup)
        metadata = move_metadata.get(normalize_key(name), {})
        move = moves.setdefault(
            move_constant,
            {
                "constant": move_constant,
                "name": name,
                "learners": {"levelUp": [], "egg": []},
                **metadata,
            },
        )
        return move

    for species in pokemon.values():
        for entry in species["levelUpLearnset"]:
            move = ensure_move(entry["moveConstant"])
            move["learners"]["levelUp"].append(
                {
                    "species": species["constant"],
                    "name": species["name"],
                    "level": entry["level"],
                }
            )
            entry["move"] = move["name"]

        for entry in species["eggMoves"]:
            move = ensure_move(entry["moveConstant"])
            move["learners"]["egg"].append({"species": species["constant"], "name": species["name"]})
            entry["move"] = move["name"]

    return sorted(moves.values(), key=lambda item: item["name"])


def collect_ability_data(pokemon: dict[str, dict], ability_metadata: dict[str, dict]) -> list[dict]:
    used_names = sorted(
        {
            name
            for species in pokemon.values()
            for name in [*species.get("abilities", []), species.get("hiddenAbility", "")]
            if name
        }
    )
    records = []
    for name in used_names:
        record = {"name": name}
        metadata = ability_metadata.get(normalize_key(name))
        if metadata:
            record.update(metadata)
        records.append(record)
    return records


def attach_locations(pokemon: dict[str, dict], location_data: dict) -> list[str]:
    pokemon_by_name = defaultdict(list)
    for species in pokemon.values():
        pokemon_by_name[normalize_key(species["name"])].append(species)
        pokemon_by_name[normalize_key(species["name"].replace("Alolan ", ""))].append(species)
        pokemon_by_name[normalize_key(species["name"].replace("Galarian ", ""))].append(species)
        pokemon_by_name[normalize_key(species["name"].replace("Hisuian ", ""))].append(species)

    unmatched = set()
    for location in location_data["locations"]:
        for method in location["methods"]:
            for species_name in method["species"]:
                if species_name.lower() == "special encounter":
                    continue
                candidates = pokemon_by_name.get(normalize_key(species_name), [])
                if not candidates:
                    unmatched.add(species_name)
                    continue
                for candidate in candidates[:1]:
                    candidate["locations"].append(
                        {
                            "location": location["name"],
                            "method": method["label"],
                            "source": method["source"],
                        }
                    )
    return sorted(unmatched)


def attach_asset_paths(pokemon: dict[str, dict], location_data: dict) -> dict[str, int]:
    sprite_index = build_asset_index(POKEMON_ASSETS_DIR)
    item_index = build_asset_index(ITEM_ASSETS_DIR)
    pokemon_with_sprites = 0
    item_icons = 0

    def decorate_item(record: dict, name_key: str = "name", *, machine_kind: str = "") -> None:
        nonlocal item_icons
        icon = resolve_item_icon(
            record.get(name_key, ""),
            item_index,
            item_type=record.get("type", ""),
            machine_kind=machine_kind,
        )
        if icon:
            record["icon"] = icon
            item_icons += 1

    for species in pokemon.values():
        species["sprite"] = resolve_sprite(species, sprite_index)
        if species["sprite"] != POKEMON_SPRITE_FALLBACK:
            pokemon_with_sprites += 1
        for held_item in species.get("heldItems", []):
            decorate_item(held_item, "item")

    for record in location_data.get("tmHm", []):
        machine_kind = "hm" if str(record.get("number", "")).upper().startswith("HM") else "tm"
        decorate_item(record, machine_kind=machine_kind)
    for record in location_data.get("megaStones", []):
        decorate_item(record)
    for record in location_data.get("zCrystals", []):
        decorate_item(record)
    for record in location_data.get("wildHeldItems", []):
        for item in record.get("items", []):
            decorate_item(item, "item")

    return {
        "pokemonSprites": pokemon_with_sprites,
        "itemIcons": item_icons,
    }


def build_data() -> dict:
    pokemon = parse_base_stats()
    learnsets = parse_learnsets()
    egg_moves = parse_egg_moves()
    evolutions = parse_evolutions()
    location_data, location_move_hints = parse_location_workbook()
    unbound_pokedex_locations = parse_unbound_pokedex_wild_locations(pokemon)
    if unbound_pokedex_locations:
        location_data["locations"] = unbound_pokedex_locations
        location_data["source"] = f'{location_data["source"]} + {UNBOUND_POKEDEX_ENCOUNTERS_FILE.name}'
    frontier_data, frontier_move_hints = parse_frontier_workbook()
    reference_metadata = load_reference_metadata()
    move_hints = {**location_move_hints, **frontier_move_hints}

    attach_source_data(pokemon, learnsets, egg_moves, evolutions)
    unmatched_location_species = attach_locations(pokemon, location_data)
    asset_counts = attach_asset_paths(pokemon, location_data)
    moves = collect_move_data(pokemon, move_hints, reference_metadata["moves"], reference_metadata["moveNames"])
    abilities = collect_ability_data(pokemon, reference_metadata["abilities"])

    pokemon_list = list(pokemon.values())
    for index, species in enumerate(pokemon_list, start=1):
        species["guideNumber"] = index

    source_counts = {
        "pokemon": len(pokemon_list),
        "pokemonWithLearnsets": sum(1 for item in pokemon_list if item["levelUpLearnset"]),
        "pokemonWithEggMoves": sum(1 for item in pokemon_list if item["eggMoves"]),
        "pokemonWithEvolutions": sum(1 for item in pokemon_list if item["evolutions"]),
        "moves": len(moves),
        "locations": len(location_data["locations"]),
        "giftStatic": len(location_data["giftStatic"]),
        "legendary": len(location_data["legendary"]),
        "trades": len(location_data["trades"]),
        "swarms": len(location_data["swarms"]),
        "wildHeldItems": len(location_data["wildHeldItems"]),
        "tmHm": len(location_data["tmHm"]),
        "megaStones": len(location_data["megaStones"]),
        "zCrystals": len(location_data["zCrystals"]),
        "frontierServices": len(frontier_data["services"]),
        "moveTutors": len(frontier_data["moveTutors"]),
        "abilities": len(abilities),
        "abilitiesWithMetadata": sum(1 for item in abilities if item.get("metadataSource")),
        "movesWithMetadata": sum(1 for item in moves if item.get("metadataSource")),
        **asset_counts,
    }

    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "game": "Pokemon Unbound",
        "version": "2.1.1.1",
        "sources": {
            "baseStats": BASE_STATS_FILE.name,
            "evolutions": EVOLUTION_FILE.name,
            "learnsets": LEARNSETS_FILE.name,
            "eggMoves": EGG_MOVES_FILE.name,
            "locations": location_data["source"],
            "frontier": frontier_data["source"],
            "referenceMetadata": reference_metadata["sources"],
        },
        "sourceCounts": source_counts,
        "pokemon": pokemon_list,
        "moves": moves,
        "abilities": abilities,
        "locations": location_data["locations"],
        "giftStatic": location_data["giftStatic"],
        "legendary": location_data["legendary"],
        "trades": location_data["trades"],
        "gameCorner": location_data["gameCorner"],
        "swarms": location_data["swarms"],
        "items": {
            "wildHeldItems": location_data["wildHeldItems"],
            "tmHm": location_data["tmHm"],
            "megaStones": location_data["megaStones"],
            "zCrystals": location_data["zCrystals"],
        },
        "frontier": frontier_data,
        "trainers": {"included": False},
        "warnings": {
            "unmatchedLocationSpecies": unmatched_location_species,
            "trainerTeams": None,
            "moveMetadata": None
            if source_counts["movesWithMetadata"]
            else "Move details are unavailable because fetched CFRU move source files were not found.",
            "abilityMetadata": None
            if source_counts["abilitiesWithMetadata"]
            else "Ability descriptions are unavailable because fetched CFRU ability source files were not found.",
            "referenceMetadata": reference_metadata["warnings"],
        },
    }


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    data = build_data()
    json_text = json.dumps(data, ensure_ascii=False, indent=2)
    (DATA_DIR / "unbound-data.json").write_text(json_text + "\n", encoding="utf-8")
    (DATA_DIR / "unbound-data.js").write_text(
        f"window.UNBOUND_GUIDE_DATA = {json_text};\n", encoding="utf-8"
    )
    print(
        json.dumps(
            {
                "output": str(DATA_DIR / "unbound-data.json"),
                "counts": data["sourceCounts"],
                "warnings": {
                    "unmatchedLocationSpecies": len(data["warnings"]["unmatchedLocationSpecies"]),
                },
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
